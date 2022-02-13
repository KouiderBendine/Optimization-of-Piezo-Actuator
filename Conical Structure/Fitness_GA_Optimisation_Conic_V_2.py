import numpy as np
import matplotlib.pyplot as plt
import math
import pathlib
import openpyxl
import random
import time
import xlsxwriter

start_time = time.time()
Population_size = int(120)
qessi = 0.019
N_Mode = 6
mutation_rate = 7
num_generations = 600
Patch_N = [2, 3, 4, 5, 6, 7, 8, 9, 10]
NFitnss = np.zeros((9, num_generations+1))
cot = 0
A_vect = np.array([155.100118564, 156.918612758, 184.685950442, 188.309246017, 195.356736455, 196.050084914])

for Patch_Number in Patch_N:
    # -------------------------------------
    Best_Fitness = 0
    X_Mut = 118

    # -------------------------------------

    Pos_Accum = np.zeros((Patch_Number, Population_size))
    Vectors = np.zeros((Patch_Number, Population_size))
    best_outputs = np.zeros((Population_size))
    Pos_vect = np.zeros((Patch_Number, Population_size))


    def Obj_Fun(Posvect, N_Mode, qessi):

        Posvect = np.sort(Posvect)
        ## Filter used to ensure no overlap of the paths
        RAND = np.array([list(range(0, 118))])
        for i in range(Patch_Number):
            for j in range(Patch_Number):
                if j != i:
                    if Posvect[int(j)] == Posvect[int(i)]:
                        arr = np.delete(RAND, np.argwhere(RAND == Posvect[j]))
                        Posvect[j] = np.random.choice(arr)
                        # print(j)
                else:
                    pass

        ## Read The Output Results
        Read_file = pathlib.Path('D:\Vibration_control_Paper\Ansys_Code\Canonical_Patch\OnePatchBmatrixConic.xlsx')
        Output = openpyxl.load_workbook(Read_file)
        Sheet = Output.active

        Sheet_Alph = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                      'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        Total_vec = np.zeros((int(Sheet.max_row), int(Sheet.max_column - 1)))
        ctr = 0
        for i in range(Sheet.max_row):
            ctr += 1
            stp = 0
            cton = 0
            for j in range(Sheet.max_column - 1):
                cton += 1
                if cton == len(Sheet_Alph):
                    stp += 1
                    cton = 0
                    if stp == 44:
                        break
                if j < len(Sheet_Alph):
                    Total_vec[i, j] = Sheet[Sheet_Alph[j] + str(ctr)].value
                else:
                    Total_vec[i, j] = Sheet[Sheet_Alph[stp - 1] + Sheet_Alph[cton] + str(ctr)].value
        ## Finsh Reading

        ## Calculate The fitness function
        LamdaS = 0
        LamdaP = 1
        for N in range(N_Mode):
            LamdaS = LamdaS + np.sum(np.square(Total_vec[N, Posvect.astype(int)])) / (4 * qessi * 2 * np.pi * A_vect[N])
            LamdaP = LamdaP * np.sum(np.square(Total_vec[N, Posvect.astype(int)])) / (4 * qessi * 2 * np.pi * A_vect[N])
        Fitn = LamdaS * (LamdaP) ** (1 / N_Mode)

        return Fitn, Posvect


    def crossover(parent_1, parent_2):
        # Get length of chromosome
        chromosome_length = len(parent_1)
        # Pick crossover point, avoding ends of chromsome
        crossover_point = random.randint(1, chromosome_length - 1)
        # Create children. np.hstack joins two arrays
        child_1 = np.hstack((parent_1[0:crossover_point],
                             parent_2[crossover_point:]))
        child_2 = np.hstack((parent_2[0:crossover_point],
                             parent_1[crossover_point:]))
        # Return children
        return child_1, child_2


    def mutate(population, X_Mut, mutation_rate):
        # Apply random mutation
        alf = int(mutation_rate * np.size(population) / 100)

        for i in range(alf):
            random_Row = random.randint(0, np.size(population, axis=0) - 1)
            random_Col = random.randint(0, np.size(population, axis=1) - 1)
            population[random_Row, random_Col] = np.random.choice(X_Mut, 1)

        # Return mutation population
        return population

    n_mutations = math.ceil((Population_size - 1) * Patch_Number * mutation_rate)

    ## Creat the starting Population
    for it in range(Population_size):
        Pos_vect[:, it] = np.random.choice(118, Patch_Number)

    ## Fitness initial evaluation
    for iter in range(Population_size):
        Ftn, Posvect = Obj_Fun(Pos_vect[:, iter], N_Mode, qessi)
        Pos_vect[:, iter] = Posvect
        best_outputs[iter] = Ftn
        Ranking = np.argsort(-1 * best_outputs)
        Vectors = np.array(Pos_vect[:, Ranking])
    best_outputs = best_outputs[Ranking]
    best_outputs = best_outputs[0:int(Population_size / 2)]

    Best_Fitness = [best_outputs[0]]
    Best_global = Best_Fitness[0]
    Best_Position = [Vectors[:, 0]]
    Vectors = Vectors[:, 0:int(Population_size / 2)]

    # score = [Best_Fitness]
    #print('Starting best score, percent target: %.10f' % Best_global)
    for generation in range(num_generations):
        #print("Generation : ", generation, "Best Fitness :", Best_Fitness[-1], "Best Fitness :", Best_Position[-1])

        # Create an empty list for new population

        new_population = []

        # Create new popualtion generating two children at a time

        for J in range(int(Population_size / 4)):
            parent_1 = Vectors[:, 2 * J]
            parent_2 = Vectors[:, 2 * J + 1]
            child_1, child_2 = crossover(parent_1, parent_2)
            new_population.append(child_1)
            new_population.append(child_2)

        # Replace the old population with the new one

        population = np.transpose(np.array(new_population))

        # Mutate the population

        population = mutate(population, X_Mut, n_mutations)

        # Evaluate the fitness function

        for iter in range(int(Population_size / 2)):
            Ftn, PosVect = Obj_Fun(population[:, iter], N_Mode, qessi)
            population[:, iter] = PosVect
            best_outputs[iter] = Ftn
            Ranking = np.argsort(-1 * best_outputs)
            Vectors = np.array(population[:, Ranking])
        best_outputs = best_outputs[Ranking]
        Best_local = best_outputs[0]

        if Best_local >= Best_global:
            Best_Position.append(Vectors[:, 0])
            Best_Fitness.append(Best_local)
            Best_global = Best_local
        else:
            Best_Position.append(Best_Position[generation])
            Best_Fitness.append(Best_Fitness[generation])

        NFitnss[cot, generation+1] = Best_Fitness[-1]
    cot = cot + 1

    print('**************')
    print(Best_Position[-1])
    # Plotting
    Length = 400
    Raidus1 = 100
    Raidus2 = 0.26 * Length + Raidus1
    Patch_width = 50
    XX = np.zeros((1, 8 * 18)).T
    YY = np.zeros((1, 8 * 18)).T
    i = 0
    AL = np.array(Best_Position[500:])
    plt.figure(Patch_Number + 1)
    for k in range(8):
        for it in range(18):
            XX[i] = k * Patch_width
            if it < 9:
                YY[i] = (Raidus2 - Raidus1) + it * Patch_width - k * (Patch_width * 0.26)
            else:
                YY[i] = (Raidus2 - Raidus1) + it * Patch_width + k * (Patch_width * 0.26)

            i = i + 1

    for it in range(8):
        i = 1
        for k in range(8):
            plt.plot(XX[int(k * 18):int(i * 18)], YY[int(k * 18):int(i * 18)], color='k')
            i = i + 1
    Xs = XX.reshape(8, 18)
    Ys = YY.reshape(8, 18)
    for k in range(18):
        ax1 = plt.plot(Xs[:, k], Ys[:, k], color='k')

    for i in range(len(Best_Position[500:])):
        for j in range(Patch_Number):
            ax2 = plt.plot(XX[int(AL[i, j]) + 1] + np.random.uniform(low=0, high=Patch_width / 9, size=(1, 1)),
                           YY[int(AL[i, j]) + 1] + np.random.uniform(low=0, high=Patch_width / 9, size=(1, 1)),
                           color='b', marker='*', linestyle='none')
    plt.xlim(-200, 1000)

    plt.show()

    # Save the best results
    workbook = xlsxwriter.Workbook(str('Opt_Posin') + str(Patch_Number) + '.xlsx')

    worksheet = workbook.add_worksheet()
    row = 0
    for col, data in enumerate(Best_Position):
        worksheet.write_column(row, col, data)

    Rw = Patch_Number + 1
    worksheet.write_row(Rw, 0, np.transpose(np.array(Best_Fitness)))

    workbook.close()
# Position

plt.figure(12)
for i in range(9):
    plt.plot((NFitnss[i, :]-np.min(NFitnss[i, :]))/(np.max(NFitnss[8, :])-np.min(NFitnss[i, :])))
plt.xlabel('Generation', fontsize=18, fontweight='bold')
plt.ylabel('Normalised Fitness', fontsize=18, fontweight='bold')
plt.legend(['2 Patch', '3 Patch', '4 Patch', '5 Patch', '6 Patch', '7 Patch', '8 Patch', '9 Patch', '10 Patch'])
plt.grid()
plt.show()
plt.ylim(-0.1, 1.1)
plt.xticks(fontsize=14)
plt.yticks(fontsize=14)
print(time.time() - start_time, "seconds")







